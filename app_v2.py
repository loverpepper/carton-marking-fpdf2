# -*- coding: utf-8 -*-
"""
Streamlit 应用 - 新版，支持多样式选择 + 批量Excel生成
"""
import streamlit as st
from PIL import Image, ImageDraw
from pathlib import Path
import io
import zipfile
import traceback
import logging
import sys
import os

# 配置日志输出到 stdout，以便 Docker logs 能够捕获
logging.basicConfig(

    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger("app")

# 屏蔽 fontTools 造成刷屏和 OOM （fpdf2 子集化字体时会狂刷 INFO 日志导致可能的卡死或假崩溃）
logging.getLogger("fontTools.subset").setLevel(logging.WARNING)

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 导入新版生成核心
from generation_core_v3 import (
    SKUConfig,
    BoxMarkGenerator,
    get_style_source_signature,
    reload_style_module,
)

# 导入所有样式以自动注册
import style_mcombo_standard
import style_simple
# 未来添加更多样式时在这里导入

import stats_db

# 设置页面配置
st.set_page_config(
    page_title="箱唛生成器",
    page_icon="📦",
    layout="wide"
)

logger.info("=== Web UI 初始化或刷新页面 ===")

# 初始化数据库
stats_db.init_db()

# ──────────────────────────────────────────────────────────────────────────────
# 辅助函数
# ──────────────────────────────────────────────────────────────────────────────

# Excel 模板列定义（顺序很重要）
TEMPLATE_COLUMNS = [
    ("sku_name",        "SKU名称",      "CA-6160-OE678BR-1",     "必填"),
    ("style_name",      "样式名称",     "mcombo_standard",        "必填。可选值见说明sheet"),
    ("length_cm",       "长度(cm)",     77.0,                     "必填"),
    ("width_cm",        "宽度(cm)",     67.5,                     "必填"),
    ("height_cm",       "高度(cm)",     47.0,                     "必填"),
    ("ppi",             "分辨率(PPI)",  150,                      "选填。默认150"),
    ("color",           "颜色",         "Beige",                  "选填"),
    ("product",         "产品名称(多行用\\n分隔)",     "Lift Recliner",          "必填"),
    ("product_fullname","产品全名(多行用\\n分隔)", "",             "选填"),
    ("size",            "尺寸标注",     "(Oversize)",             "选填"),
    ("gw_value",        "毛重(lbs)",    106.9,                    "选填"),
    ("nw_value",        "净重(lbs)",    94.4,                     "选填"),
    ("dim_l_in",        "箱长(英寸)",   30.3,                     "选填"),
    ("dim_w_in",        "箱宽(英寸)",   26.6,                     "选填"),
    ("dim_h_in",        "箱高(英寸)",   18.5,                     "选填"),
    ("sn_code",         "SN条形码",     "08429381073953",         "选填"),
    ("origin_text",     "原产地文字",   "MADE IN CHINA",          "选填。默认MADE IN CHINA"),
    ("total_boxes",     "总箱数",       3,                        "必填"),
    ("current_box",     "当前箱号",     1,                        "必填"),
    ("sponge_verified",           "海绵认证",              "否",                                              "选填。填 是 或 否"),
    ("country",                   "国家/地区",             "GE",                                              "选填。新市场样式用，可选: UK / FR / GE，留空忽略"),
    ("company_name",              "公司名称",              "NEWACME LLC",                                     "选填。新市场样式用"),
    ("contact_info",              "联系方式",              "www.mcombo.com / sale_uk@newacmellc.com",         "选填。新市场样式用"),
    ("show_fsc",                  "显示FSC标志",           "否",                                              "选填。填 是 或 否"),
    ("legal_3_2",                 "法律条款3_2",           0,                                                 "选填。0或1"),
    ("legal_3_3",                 "法律条款3_3",           0,                                                 "选填。0或1"),
    ("legal_3_4",                 "法律条款3_4",           0,                                                 "选填。0或1"),
    ("legal_3_5",                 "法律条款3_5",           1,                                                 "选填。0或1"),
    ("legal_3_6",                 "法律条款3_6",           1,                                                 "选填。0或1"),
    ("legal_product_name",        "法律标签-产品名",       "Lift Recliner",                                   "选填。新市场专用"),
    ("legal_model",               "法律标签-型号",         "GE-6160-LC001BG-1",                               "选填。新市场专用"),
    ("legal_batch_number",        "法律标签-批次号",       "08429381118265",                                  "选填。新市场专用"),
    ("legal_country_origin",      "法律标签-原产国",       "Made in China",                                   "选填。新市场专用"),
    ("legal_manufacturer",        "法律标签-制造商",       "TAIYUAN TEMARU ELECTRONICS TECHNOLOGY CO,.LTD",   "选填。新市场专用"),
    ("legal_manufacturer_address","法律标签-制造商地址",   "NO.201-01, Zhongchuang Space...",                 "选填。新市场专用"),
    ("legal_manufacturer_email",  "法律标签-制造商邮箱",   "cs@elegue.com",                                   "选填。新市场专用"),
]

COL_KEYS = [c[0] for c in TEMPLATE_COLUMNS]


PDF_LABEL_TRANSLATION = str.maketrans({
    "（": "(",
    "）": ")",
    "［": "[",
    "］": "]",
    "【": "[",
    "】": "]",
    "－": "-",
    "–": "-",
    "—": "-",
    "／": "/",
    "：": ":",
    "，": ",",
    "　": " ",
})


def normalize_pdf_label(value) -> str:
    """Normalize common full-width punctuation before drawing with Latin fonts."""
    return str(value).translate(PDF_LABEL_TRANSLATION).strip()


@st.cache_data(show_spinner=False)
def get_available_styles_cached() -> list:
    """Return style metadata without repeating registry work on every rerun."""
    return BoxMarkGenerator.list_available_styles()


def ensure_style_code_current(style_name: str) -> None:
    """
    在生成动作前检查目标样式文件是否被本地修改。

    Streamlit  rerun 不会自动 reload 已导入的 Python 模块；这里仅在样式
    源文件 mtime/size 变化时重载当前样式模块，并丢弃该样式的生成器缓存。
    """
    if "_style_source_signatures" not in st.session_state:
        st.session_state._style_source_signatures = {}

    signatures = st.session_state._style_source_signatures
    current_signature = get_style_source_signature(style_name)
    previous_signature = signatures.get(style_name)
    if previous_signature == current_signature:
        return

    if previous_signature is None:
        logger.info(f"首次生成前刷新样式模块: {style_name}")
    else:
        logger.info(f"检测到样式代码变更，重载样式模块: {style_name}")

    current_signature = reload_style_module(style_name)
    signatures[style_name] = current_signature

    if "_boxmark_generator_cache" in st.session_state:
        cache = st.session_state._boxmark_generator_cache
        for cache_key in list(cache.keys()):
            if cache_key[0] == style_name:
                del cache[cache_key]

    get_available_styles_cached.clear()


def get_boxmark_generator(style_name: str, ppi: int, refresh_style_code: bool = False) -> BoxMarkGenerator:
    """Reuse loaded style resources within the current Streamlit session."""
    if refresh_style_code:
        ensure_style_code_current(style_name)

    cache_key = (style_name, int(ppi))
    if "_boxmark_generator_cache" not in st.session_state:
        st.session_state._boxmark_generator_cache = {}
    cache = st.session_state._boxmark_generator_cache
    if cache_key not in cache:
        cache[cache_key] = BoxMarkGenerator(
            base_dir=Path(__file__).parent,
            style_name=style_name,
            ppi=int(ppi),
        )
    return cache[cache_key]


@st.cache_data(show_spinner=False)
def build_excel_template() -> bytes:
    """生成带样式的Excel模板，返回字节流"""
    wb = Workbook()
    ws = wb.active
    ws.title = "批量箱唛数据"

    header_fill  = PatternFill("solid", fgColor="1F4E79")
    example_fill = PatternFill("solid", fgColor="D9E1F2")
    note_fill    = PatternFill("solid", fgColor="FFF2CC")
    white_font   = Font(bold=True, color="FFFFFF", name="微软雅黑", size=10)
    normal_font  = Font(name="微软雅黑", size=10)
    center       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # 第1行：字段名（英文key）
    # 第2行：中文说明
    # 第3行：示例数据
    # 第4行：备注
    ws.append([c[0] for c in TEMPLATE_COLUMNS])   # row1
    ws.append([c[1] for c in TEMPLATE_COLUMNS])   # row2
    ws.append([c[2] for c in TEMPLATE_COLUMNS])   # row3
    ws.append([c[3] for c in TEMPLATE_COLUMNS])   # row4

    for col_idx, col_def in enumerate(TEMPLATE_COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(16, len(str(col_def[1])) * 2.2)

        for row_idx in range(1, 5):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = center
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = white_font
            elif row_idx == 2:
                cell.fill = header_fill
                cell.font = Font(bold=False, color="DDDDDD", name="微软雅黑", size=9)
            elif row_idx == 3:
                cell.fill = example_fill
                cell.font = normal_font
            elif row_idx == 4:
                cell.fill = note_fill
                cell.font = Font(italic=True, color="7F7F7F", name="微软雅黑", size=9)

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 30
    ws.freeze_panes = "A5"

    # 样式说明 sheet
    ws2 = wb.create_sheet("样式说明")
    ws2.append(["样式名称(style_name)", "说明"])
    available_styles = get_available_styles_cached()
    for s in available_styles:
        ws2.append([s['name'], s['description']])
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 50

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def row_to_skuconfig(row: dict, base_dir) -> SKUConfig:
    """将 DataFrame 行（dict）转换为 SKUConfig 对象"""

    def get(key, default=None):
        v = row.get(key, default)
        if pd.isna(v) if not isinstance(v, str) else (v.strip() == ""):
            return default
        return v

    sku_name   = str(get("sku_name", "unknown_sku")).strip()
    style_name = str(get("style_name", "mcombo_standard")).strip()
    length_cm  = float(get("length_cm", 0))
    width_cm   = float(get("width_cm", 0))
    height_cm  = float(get("height_cm", 0))
    ppi        = int(float(get("ppi", 150)))

    style_params = {}

    color = get("color")
    if color is not None:
        style_params["color"] = str(color).strip()

    product = get("product")
    if product is not None:
        style_params["product"] = str(product).strip()

    product_fullname = get("product_fullname")
    if product_fullname is not None:
        style_params["product_fullname"] = str(product_fullname).replace("\\n", "\n").strip()

    size = get("size")
    if size is not None:
        style_params["size"] = normalize_pdf_label(size)

    sponge_raw = get("sponge_verified", "否")
    style_params["sponge_verified"] = str(sponge_raw).strip() in ("是", "True", "true", "1", "yes", "Yes")

    # 组装 side_text 字典
    gw   = get("gw_value")
    nw   = get("nw_value")
    sn   = get("sn_code")
    orig = get("origin_text", "MADE IN CHINA")
    dl   = get("dim_l_in")
    dw   = get("dim_w_in")
    dh   = get("dim_h_in")

    side_text = {}
    if gw is not None:
        side_text["gw_value"] = float(gw)
    if nw is not None:
        side_text["nw_value"] = float(nw)
    if sn is not None:
        side_text["sn_code"] = str(sn).strip()
    if orig is not None:
        side_text["origin_text"] = str(orig).strip()
    if dl is not None and dw is not None and dh is not None:
        side_text["dimention_text"] = f"BOX SIZE: {float(dl):.1f}'' x {float(dw):.1f}'' x {float(dh):.1f}''"
    if side_text:
        style_params["side_text"] = side_text

    # 组装 box_number 字典
    total_boxes  = get("total_boxes")
    current_box  = get("current_box")
    if total_boxes is not None and current_box is not None:
        style_params["box_number"] = {
            "total_boxes": int(float(total_boxes)),
            "current_box":  int(float(current_box))
        }

    # ── 新市场专用参数 ──
    country      = get("country")  # None 表示不启用国家开关
    company_name = str(get("company_name", "NEWACME LLC")).strip()
    contact_info = str(get("contact_info", "www.mcombo.com / sale_uk@newacmellc.com")).strip()

    show_fsc    = str(get("show_fsc",    "否")).strip() in ("是", "True", "true", "1", "yes", "Yes")
    legal_3_2 = int(float(get("legal_3_2", 0) or 0))
    legal_3_3 = int(float(get("legal_3_3", 0) or 0))
    legal_3_4 = int(float(get("legal_3_4", 0) or 0))
    legal_3_5 = int(float(get("legal_3_5", 0) or 0))
    legal_3_6 = int(float(get("legal_3_6", 0) or 0))

    # 组装 legal_data 字典
    lp     = get("legal_product_name")
    lm     = get("legal_model")
    lb     = get("legal_batch_number")
    lco    = get("legal_country_origin")
    lmfr   = get("legal_manufacturer")
    lmaddr = get("legal_manufacturer_address")
    lmeml  = get("legal_manufacturer_email")
    legal_data = None
    if any(v is not None for v in [lp, lm, lb, lco, lmfr, lmaddr, lmeml]):
        legal_data = {
            "Product Name":         lp     or "",
            "Model":                lm     or "",
            "Batch Number":         lb     or "",
            "Country Origin":       lco    or "",
            "Manufacturer":         lmfr   or "",
            "Manufacturer Address": lmaddr or "",
            "Manufacturer E-mail":  lmeml  or "",
        }

    return SKUConfig(
        sku_name=sku_name,
        length_cm=length_cm,
        width_cm=width_cm,
        height_cm=height_cm,
        style_name=style_name,
        ppi=ppi,
        company_name=company_name,
        contact_info=contact_info,
        legal_data=legal_data,
        legal_3_2=legal_3_2,
        legal_3_3=legal_3_3,
        legal_3_4=legal_3_4,
        legal_3_5=legal_3_5,
        legal_3_6=legal_3_6,
        show_fsc=show_fsc,
        country=country if country else None,
        **style_params
    )

# ──────────────────────────────────────────────────────────────────────────────
# 初始化 session state
# ──────────────────────────────────────────────────────────────────────────────
if 'generated_image' not in st.session_state:
    st.session_state.generated_image = None
if 'pdf_bytes' not in st.session_state:
    st.session_state.pdf_bytes = None
if 'batch_zip_bytes' not in st.session_state:
    st.session_state.batch_zip_bytes = None
if 'batch_results' not in st.session_state:
    st.session_state.batch_results = []   # list of (sku_name, ok, msg)
if 'last_gen_info' not in st.session_state:
    st.session_state.last_gen_info = None

# 页面标题
st.title("📦 Mcombo·Barberpub·Exacme·新市场 箱唛生成器 V4")
st.caption("🎨 支持多样式切换 · 支持批量Excel生成")


# ──────────────────────────────────────────────────────────────────────────────
# Tab 切换
# ──────────────────────────────────────────────────────────────────────────────
tab_single, tab_batch, tab_stats = st.tabs(["🖊️ 单个生成", "📊 批量生成（Excel）", "📈 数据看板🔒"])

# 使用自定义 CSS 将第三个 Tab（数据看板）强制推到最右侧
st.markdown("""
<style>
    div[data-baseweb="tab-list"] {
        display: flex;
        width: 100%;
    }
    div[data-baseweb="tab"]:nth-child(3) {
        margin-left: auto !important;
    }
</style>
""", unsafe_allow_html=True)

# ──────────────────────────────────────────────────────────────────────────────
# Tab 1：单个生成（原有功能）
# ──────────────────────────────────────────────────────────────────────────────
with tab_single:

    # 获取所有可用样式
    available_styles = get_available_styles_cached()
    style_names = [s['name'] for s in available_styles]
    style_descriptions = {s['name']: s['description'] for s in available_styles}

    # 样式选择器
    st.header("🎨 样式选择")
    selected_style = st.selectbox(
        "选择箱唛样式",
        options=style_names,
        format_func=lambda x: f"{x} - {style_descriptions[x]}",
        index=0
    )
    st.info(f"当前样式: **{selected_style}** - {style_descriptions[selected_style]}")

    current_style_info = next((s for s in available_styles if s['name'] == selected_style), None)
    required_params = current_style_info['required_params'] if current_style_info else []

    st.markdown("---")

    # ══════════════════════════════════════════════════════════════════
    # 区块一：基本信息（通栏，横向平铺）
    # ══════════════════════════════════════════════════════════════════
    st.header("📝 基本信息")

    # 行 1：SKU 名称 × 4 格  |  PPI × 1 格  |  总箱数 × 1 格  |  当前箱号 × 1 格
    bi1, bi2, bi3, bi4 = st.columns([4, 1, 1, 1])
    with bi1:
        sku_name = st.text_input("SKU 名称", value="CA-6160-OE678BR-1",
                                 help="例如: CA-6160-OE678BR-1")
    with bi2:
        ppi = st.selectbox("分辨率 PPI", options=[72, 150, 300], index=1,
                           help="150适合预览，300适合印刷")
    with bi3:
        total_boxes = st.number_input("总箱数", min_value=1, value=3, step=1)
    with bi4:
        current_box = st.number_input("当前箱号", min_value=1, value=1, step=1)

    # 行 2：长 / 宽 / 高 (cm) × 3 格  |  自动换算英寸（只读提示）× 1 格
    st.caption("📐 箱子尺寸 (cm)   ← 输入后英寸自动换算")
    bi5, bi6, bi7, bi8 = st.columns([1, 1, 1, 2])
    with bi5:
        length_cm = st.number_input("长度 L (cm)", min_value=1.0, value=77.0, step=0.5)
    with bi6:
        width_cm  = st.number_input("宽度 W (cm)", min_value=1.0, value=67.5, step=0.5)
    with bi7:
        height_cm = st.number_input("高度 H (cm)", min_value=1.0, value=47.0, step=0.5)

    # 自动换算英寸（用于侧唛 dimention_text，无需手动输入）
    _l_in = round(length_cm / 2.54, 1)
    _w_in = round(width_cm  / 2.54, 1)
    _h_in = round(height_cm / 2.54, 1)
    with bi8:
        st.metric("换算英寸 (自动)",
                  f"{_l_in}'' × {_w_in}'' × {_h_in}''",
                  help="由厘米自动计算，无需手动填写")

    st.markdown("---")

    # ══════════════════════════════════════════════════════════════════
    # 区块二：样式参数（通栏，三列卡片式网格）
    # ══════════════════════════════════════════════════════════════════
    st.header("📋 样式参数")
    style_params = {}

    sp_col1, sp_col2, sp_col3 = st.columns([1, 1, 1])

    # ── 列 1：产品信息 ──
    with sp_col1:
        st.subheader("🏷️ 产品信息")
        color = st.text_input("颜色 color", value="Beige",
                              help="例如: Beige, Brown, Seafoam Green")
        size  = st.text_input("尺寸标注 size", value="(Oversize)",
                              help="例如: (Oversize), (Medium-Wide)")
        product = st.text_input("产品名称 product 【多行用 \\n 分隔】", value="Lift Recliner",
                                help="例如: Lift Recliner, Rectangle Trampoline")
        product_fullname = st.text_input(
            "产品全名 product_fullname （只适用Exacme 品牌）【多行用 \\n 分隔】", value="",
            help="多行用 \\n 分隔，部分样式专用，留空忽略"
        )
        if color:   style_params['color'] = color
        if product: style_params['product'] = product
        if product_fullname.strip():
            style_params['product_fullname'] = product_fullname.replace("\\n", "\n")
        if size:    style_params['size'] = normalize_pdf_label(size)

    # ── 列 2：侧唛信息 ──
    with sp_col2:
        st.subheader("📦 侧唛信息")
        gross_weight = st.number_input("毛重 gw (lbs)", min_value=0.0, value=106.9, step=0.1)
        net_weight   = st.number_input("净重 nw (lbs)", min_value=0.0, value=94.4,  step=0.1)
        sn_code      = st.text_input("SN 条形码", value="08429381073953",
                                     help="条形码序列号")
        origin_text  = st.text_input("原产地文字", value="MADE IN CHINA")
        st.caption(f"📐 箱尺寸英寸（自动）: {_l_in}'' × {_w_in}'' × {_h_in}''")
        style_params['side_text'] = {
            'gw_value':       gross_weight,
            'nw_value':       net_weight,
            'dimention_text': f"BOX SIZE: {_l_in}'' x {_w_in}'' x {_h_in}''",
            'sn_code':        sn_code,
            'origin_text':    origin_text,
        }

    # ── 列 3：其他选项 ──
    with sp_col3:
        st.subheader("⚙️ 其他选项")
        sponge_verified = st.selectbox("海绵认证 sponge_verified",
                                       options=["否", "是"], index=0) == "是"
        style_params['sponge_verified'] = sponge_verified

        st.markdown("**线描图**（Barberpub 对开盖和 Exacme双圈埋地蹦床专用背景）")
        line_drawing_file = st.file_uploader(
            "上传线描图 (PNG / JPG / PDF矢量)",
            type=["png", "jpg", "jpeg", "pdf"],
            key="line_drawing_uploader",
            help="Barberpub 对开盖和 Exacme双圈埋地蹦床样式专用背景图，支持PDF矢量图"
        )
        if line_drawing_file is not None:
            import tempfile
            suffix = Path(line_drawing_file.name).suffix
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
            tmp.write(line_drawing_file.read())
            tmp.close()
            style_params['img_line_drawing'] = Path(tmp.name)
            if suffix.lower() == '.pdf':
                st.info("已上传 PDF 矢量线描图，预览不可用，生成时将以矢量方式嵌入。")
            else:
                st.image(line_drawing_file, caption="线描图预览", width='stretch')

    # ── 新市场专用参数（折叠区块）──
    st.markdown("---")
    with st.expander("🌍 新市场专用参数（New Market 样式适用，其他样式可忽略）"):
        nm_col1, nm_col2, nm_col3 = st.columns([1, 1, 1])

        with nm_col1:
            st.subheader("🏢 公司信息")
            nm_country      = st.selectbox("市场/国家 country", options=["", "GE", "UK", "FR"], index=0,
                                           help="选择目标市场，留空则不启用国家开关")
            nm_company_name = st.text_input("公司名称 company_name", value="NEWACME LLC")
            nm_contact_info = st.text_input("联系方式 contact_info",
                                            value="www.mcombo.com / sale_uk@newacmellc.com")
            nm_show_fsc    = st.selectbox("显示FSC标志 show_fsc",    options=["否", "是"], index=0) == "是"

        with nm_col2:
            st.subheader("📋 法律条款开关")
            nm_legal_3_2 = int(st.checkbox("legal_3_2", value=False))
            nm_legal_3_3 = int(st.checkbox("legal_3_3", value=False))
            nm_legal_3_4 = int(st.checkbox("legal_3_4", value=False))
            nm_legal_3_5 = int(st.checkbox("legal_3_5", value=True))
            nm_legal_3_6 = int(st.checkbox("legal_3_6", value=True))

        with nm_col3:
            st.subheader("🏷️ 法律标签信息 legal_data")
            nm_lp    = st.text_input("产品名 Product Name",           value="Lift Recliner")
            nm_lm    = st.text_input("型号 Model",                     value="")
            nm_lb    = st.text_input("批次号 Batch Number",            value="")
            nm_lco   = st.text_input("原产国 Country Origin",          value="Made in China")
            nm_lmfr  = st.text_input("制造商 Manufacturer",            value="")
            nm_laddr = st.text_input("制造商地址 Manufacturer Address", value="")
            nm_leml  = st.text_input("制造商邮箱 Manufacturer E-mail",  value="")

    nm_legal_data = {
        "Product Name":         nm_lp,
        "Model":                nm_lm,
        "Batch Number":         nm_lb,
        "Country Origin":       nm_lco,
        "Manufacturer":         nm_lmfr,
        "Manufacturer Address": nm_laddr,
        "Manufacturer E-mail":  nm_leml,
    }

    # 箱号写入 style_params
    style_params['box_number'] = {
        'total_boxes': int(total_boxes),
        'current_box':  int(current_box)
    }

    # 生成按钮区域
    st.markdown("---")
    col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 2])

    with col_btn1:
        generate_preview = st.button("🖼️ 生成预览", type="primary")

    with col_btn2:
        current_sku = sku_name.strip() if sku_name and sku_name.strip() else "carton_marking"
        if st.session_state.pdf_bytes:
            st.download_button(
                label="📥 下载 PDF",
                data=st.session_state.pdf_bytes,
                file_name=f"{current_sku}.pdf",
                mime="application/pdf"
            )
        else:
            st.button("📥 下载 PDF (请先生成预览)", disabled=True)

    # 生成逻辑
    if generate_preview:
        with st.spinner(f"正在使用 {selected_style} 样式生成箱唛..."):
            try:
                test_sku = SKUConfig(
                    sku_name=sku_name,
                    length_cm=length_cm,
                    width_cm=width_cm,
                    height_cm=height_cm,
                    style_name=selected_style,
                    ppi=ppi,
                    company_name=nm_company_name,
                    contact_info=nm_contact_info,
                    legal_data=nm_legal_data,
                    legal_3_2=nm_legal_3_2,
                    legal_3_3=nm_legal_3_3,
                    legal_3_4=nm_legal_3_4,
                    legal_3_5=nm_legal_3_5,
                    legal_3_6=nm_legal_3_6,
                    show_fsc=nm_show_fsc,
                    country=nm_country if nm_country else None,
                    **style_params
                )
                generator = get_boxmark_generator(selected_style, ppi, refresh_style_code=True)

                # 生成 PDF 字节（fpdf2 直接输出，矢量清晰）
                pdf_bytes_data = generator.generate_pdf_bytes(test_sku)
                st.session_state.pdf_bytes = pdf_bytes_data

                # 用 PyMuPDF 将 PDF 直接极低内存安全压缩成小图
                import fitz
                doc = fitz.open(stream=pdf_bytes_data, filetype="pdf")
                page = doc[0]

                # PDF的原始 pt 尺寸 (72 DPI)
                pdf_w_pt = page.rect.width
                pdf_h_pt = page.rect.height

                # 计算用户请求原本会产生的真实大图尺寸 (按用户指定的 PPI)
                true_w_px = int(pdf_w_pt * ppi / 72.0)
                true_h_px = int(pdf_h_pt * ppi / 72.0)

                # ⚠️ 内存安全保护机制：如果原始像素尺寸超过 2000，则渲染时直接使用缩小比例
                # 这样可以彻底避免创建一张 3.5GB 甚至 10GB 的超级缓存光栅图导致 OOM 崩溃！
                max_preview_w = 2000
                render_zoom = ppi / 72.0
                if true_w_px > max_preview_w:
                    render_zoom = max_preview_w / pdf_w_pt
                
                mat = fitz.Matrix(render_zoom, render_zoom)
                pix = page.get_pixmap(matrix=mat, alpha=False)  # 用 RGB 降低开销，避免巨大的 RGBA 通道
                preview_image = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                
                doc.close()

                st.session_state.generated_image = preview_image
                st.session_state.last_gen_info = (style_descriptions[selected_style], true_w_px, true_h_px, ppi)
                logger.info(f"单张生成成功: SKU={sku_name}, 样式={selected_style}")
                
                # 记录统计 (同时记录产品名称)
                _prod_name = style_params.get('product', '')
                stats_db.log_success(selected_style, current_sku, "单张(含预览)", _prod_name)

            except Exception as e:
                logger.error(f"单张生成出错: SKU={sku_name}, error: {str(e)}", exc_info=True)
                st.error(f"❌ 生成失败: {str(e)}")
                st.code(traceback.format_exc())

    # 预览区
    if st.session_state.generated_image:
        st.markdown("---")
        if 'last_gen_info' in st.session_state:
            _sty, _tw, _th, _ppi = st.session_state.last_gen_info
            st.success(f"✅ 箱唛生成成功！（样式: {_sty}）")
            st.info(f"📐 PDF 尺寸: {_tw}x{_th}px | 🎨 分辨率: {_ppi} PPI")
        st.header("🖼️ 预览")
        st.image(st.session_state.generated_image, width='stretch')
        if st.session_state.pdf_bytes:
            st.info("💡 提示：预览图已生成，点击上方'下载 PDF'按钮保存文件")

    # 底部说明
    st.markdown("---")
    available_styles_list = get_available_styles_cached()
    st.markdown("### 🎨 当前可用样式")
    for style_info in available_styles_list:
        st.markdown(f"- **{style_info['name']}**: {style_info['description']}")
        if style_info['required_params']:
            st.markdown(f"  - 必需参数: `{', '.join(style_info['required_params'])}`")
    st.markdown("""
### 🌐 局域网访问
哪里都能访问了，快分享给需要的同事吧！如果有任何问题或建议，欢迎随时联系数据组。我们会持续优化功能和用户体验！🚀
""")


# ──────────────────────────────────────────────────────────────────────────────
# Tab 2：批量生成（Excel）
# ──────────────────────────────────────────────────────────────────────────────
with tab_batch:
    st.header("📊 批量箱唛生成")
    st.markdown("""
    **使用步骤**
    1. 下载 Excel 模板，按表头填写每个 SKU 的参数（第3行为示例行，可直接编辑）
    2. 上传填好的 Excel 文件
    3. 预览数据无误后点击 **批量生成** 按钮
    4. 生成完成后一键下载包含所有 PDF 的 ZIP 包
    """)

    # ── 模板下载 ──
    col_tpl1, col_tpl2 = st.columns([1, 3])
    with col_tpl1:
        tpl_bytes = build_excel_template()
        st.download_button(
            label="📥 下载 Excel 模板",
            data=tpl_bytes,
            file_name="箱唛批量生成模板.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="secondary"
        )
    with col_tpl2:
        st.info("模板第1行=字段名（勿修改），第2行=中文说明，第3行=示例数据，从**第5行**开始填写实际数据。")

    st.markdown("---")

    # ── 文件上传 ──
    uploaded_file = st.file_uploader(
        "上传填好的 Excel 文件（.xlsx / .xls）",
        type=["xlsx", "xls"],
        help="请按模板格式填写，第1行为字段名（英文key），从第5行起为数据行。"
    )

    if uploaded_file is not None:
        try:
            # 引入缓存机制：避免Streamlit在每次键盘输入时都去读取解析巨大的Excel文件
            @st.cache_data(show_spinner=False)
            def load_and_parse_excel(file_bytes: bytes) -> pd.DataFrame:
                df = pd.read_excel(io.BytesIO(file_bytes), header=0, skiprows=[1, 2, 3], dtype=str)
                df.dropna(how="all", inplace=True)
                return df

            # 读取 Excel：第1行为字段名(header)，跳过第2-4行（中文说明/示例/备注），第5行起为数据
            df_raw = load_and_parse_excel(uploaded_file.getvalue()).copy()

            if df_raw.empty:
                st.warning("⚠️ 上传的文件没有有效数据行（跳过了说明行后为空）。请从第5行起填写数据。")
            else:
                st.subheader(f"📋 读取到 {len(df_raw)} 条 SKU 数据（可直接在表格内编辑）")

                _style_options = [s['name'] for s in get_available_styles_cached()]
                display_cols = [c for c in COL_KEYS if c in df_raw.columns]
                edit_df = df_raw[display_cols].copy() if display_cols else df_raw.copy()

                # 为 style_name 列配置下拉菜单
                col_cfg = {}
                if 'style_name' in edit_df.columns:
                    col_cfg['style_name'] = st.column_config.SelectboxColumn(
                        label="样式名称",
                        options=_style_options,
                        required=True,
                        help="选择要使用的箱唛样式"
                    )

                df_edited = st.data_editor(
                    edit_df,
                    column_config=col_cfg,
                    width='stretch',
                    height=280,
                    num_rows="dynamic",
                    key="batch_editor"
                )
                # 将编辑后的数据同步回 df_raw（仅 display_cols 部分）
                for c in display_cols:
                    df_raw[c] = df_edited[c].values

                st.markdown("---")

                # ── 批量生成按钮 ──
                batch_ppi = st.selectbox(
                    "批量生成分辨率 (PPI)",
                    options=[72, 150, 300], index=1,
                    help="覆盖Excel中ppi列的值（如Excel中有值则以Excel为准）",
                    key="batch_ppi"
                )

                col_gen, col_dl = st.columns([1, 1])
                with col_gen:
                    do_batch = st.button("🚀 批量生成全部箱唛", type="primary")

                # ── 执行批量生成 ──
                if do_batch:
                    logger.info(f"开始批量生成任务，共 {len(df_raw)} 条数据进入队列。")
                    base_dir = Path(__file__).parent
                    zip_buffer = io.BytesIO()
                    results = []

                    progress_bar = st.progress(0, text="准备生成...")
                    status_area  = st.empty()

                    with zipfile.ZipFile(zip_buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
                        total = len(df_raw)
                        for i, (_, row) in enumerate(df_raw.iterrows()):
                            row_dict = row.to_dict()
                            # 如果 ppi 列为空，使用界面选择的 ppi
                            if pd.isna(row_dict.get("ppi", None)) or str(row_dict.get("ppi", "")).strip() == "":
                                row_dict["ppi"] = str(batch_ppi)

                            sku_label = str(row_dict.get("sku_name", f"row_{i+1}")).strip()
                            status_area.info(f"正在生成 [{i+1}/{total}]：{sku_label} ...")

                            try:
                                sku_cfg  = row_to_skuconfig(row_dict, base_dir)
                                gen = get_boxmark_generator(sku_cfg.style_name, sku_cfg.ppi, refresh_style_code=True)
                                pdf_bytes_item = gen.generate_pdf_bytes(sku_cfg)

                                # 文件名去除非法字符
                                safe_name = "".join(
                                    c if c.isalnum() or c in "-_. " else "_"
                                    for c in sku_label
                                ).strip()
                                zf.writestr(f"{safe_name}.pdf", pdf_bytes_item)
                                results.append((sku_label, True, "✅ 成功"))
                                logger.info(f"批量任务成功 [{i+1}/{total}]: SKU={sku_label}")
                                
                                # 将该条成功记录写入统计库，归类为批量，并提取产品名称用于记录
                                _prod_name = str(row_dict.get("product", "")).strip()
                                stats_db.log_success(sku_cfg.style_name, sku_label, "批量ZIP", _prod_name)

                            except Exception as e:
                                logger.error(f"批量生成失败 [{i+1}/{total}]: SKU={sku_label}, error: {str(e)}", exc_info=True)
                                results.append((sku_label, False, f"❌ 失败：{str(e)}"))

                            progress_bar.progress((i + 1) / total,
                                                  text=f"[{i+1}/{total}] {sku_label}")

                    st.session_state.batch_zip_bytes = zip_buffer.getvalue()
                    st.session_state.batch_results   = results
                    status_area.empty()
                    progress_bar.empty()

                    ok_count  = sum(1 for _, ok, _ in results if ok)
                    fail_count = len(results) - ok_count
                    if fail_count == 0:
                        st.success(f"🎉 全部 {ok_count} 个箱唛生成成功！")
                    else:
                        st.warning(f"⚠️ 完成：{ok_count} 个成功，{fail_count} 个失败，详见下方结果。")

                # ── 下载按钮（始终显示已生成的结果）──
                if st.session_state.batch_zip_bytes:
                    with col_dl:
                        st.download_button(
                            label=f"📦 下载全部 PDF（ZIP）",
                            data=st.session_state.batch_zip_bytes,
                            file_name="箱唛批量生成结果.zip",
                            mime="application/zip",
                            type="primary"
                        )

                # ── 结果明细 ──
                if st.session_state.batch_results:
                    st.markdown("#### 生成结果明细")
                    result_data = [
                        {"SKU": sku, "状态": msg}
                        for sku, ok, msg in st.session_state.batch_results
                    ]
                    st.dataframe(pd.DataFrame(result_data), width='stretch')

        except Exception as e:
            logger.error(f"读取或解析 Excel 文件出错: {str(e)}", exc_info=True)
            st.error(f"❌ 读取 Excel 出错：{str(e)}")
            st.code(traceback.format_exc())
    else:
        st.info("👆 请先上传 Excel 文件，或点击上方按钮下载模板。")


# ──────────────────────────────────────────────────────────────────────────────
# Tab 3：数据看板 (需密码)
# ──────────────────────────────────────────────────────────────────────────────
with tab_stats:
    st.header("📈 后台数据统计 (管理员)")
    
    # 从环境变量读取设定的密码，如果没有设定就使用默认密码 "admin2026"
    import os
    ADMIN_PWD = os.getenv("STATS_PASSWORD", "admin2026")
    
    if "admin_auth" not in st.session_state:
        st.session_state.admin_auth = False

    if not st.session_state.admin_auth:
        st.info("🔒 为了数据隐私安全，请先输入管理员密码进行身份认证。")
        col_pwd, col_btn = st.columns([3, 1])
        with col_pwd:
            pwd_input = st.text_input("请输入访问密码", type="password", key="stats_pwd_input")
        with col_btn:
            st.write("") # 占位
            st.write("") # 向下对齐
            if st.button("🔑 登录并查看"):
                if pwd_input == ADMIN_PWD:
                    st.session_state.admin_auth = True
                    st.rerun()
                else:
                    st.error("❌ 密码错误！")
    else:
        col_s1, col_s2 = st.columns([4, 1])
        with col_s1:
            st.success("✅ 认证成功！欢迎进入管理员数据看板。")
        with col_s2:
            if st.button("🚪 退出登录"):
                st.session_state.admin_auth = False
                st.rerun()
        
        # 分为两个显示块，最近几天的统计 / 历史总排名
        st.markdown("---")
        st.subheader("🗓️ 每日生成统计 (柱状图)")
        
        df_daily = stats_db.get_daily_stats()
        if not df_daily.empty:
            df_daily['生成日期'] = pd.to_datetime(df_daily['生成日期']).dt.date
            
            # --- 筛选器 ---
            f_col1, f_col2 = st.columns(2)
            with f_col1:
                min_date = df_daily['生成日期'].min()
                max_date = df_daily['生成日期'].max()
                sel_dates = st.date_input("📅 选择日期范围", value=(min_date, max_date), min_value=min_date, max_value=max_date)
            
            with f_col2:
                all_styles = df_daily['选择样式'].unique().tolist()
                sel_styles = st.multiselect("🎨 筛选目标样式", options=all_styles, default=all_styles)
            
            # --- 数据过滤 ---
            mask = pd.Series(True, index=df_daily.index)
            if len(sel_dates) == 2:
                mask &= (df_daily['生成日期'] >= sel_dates[0]) & (df_daily['生成日期'] <= sel_dates[1])
            elif len(sel_dates) == 1:
                # 只有选择了一天的情况
                mask &= (df_daily['生成日期'] == sel_dates[0])
            
            if sel_styles:
                mask &= df_daily['选择样式'].isin(sel_styles)
                
            filtered_df = df_daily[mask]
            
            # --- 显示图表和表格 ---
            if not filtered_df.empty:
                # 透视表：横坐标为日期，纵坐标为件数，按选择样式分组（柱子）(改用 .size() 统计记录条数)
                chart_data = filtered_df.groupby(['生成日期', '选择样式']).size().unstack(fill_value=0)
                
                st.bar_chart(chart_data)
                
                with st.expander("📝 查看数据明细表 (含具体SKU与名称)"):
                    st.dataframe(filtered_df.sort_values(by=['生成日期', '时间'], ascending=[False, False]), width='stretch', hide_index=True)
            else:
                st.warning("⚠️ 该日期范围或筛选条件下目前没有记录。")
        else:
            st.info("目前还没有任何生成记录哦。")
            
        st.markdown("---")
        st.subheader("🏆 全生命周期样式热度排行榜")
        df_overall = stats_db.get_overall_stats()
        if not df_overall.empty:
            st.dataframe(df_overall, width='stretch', hide_index=True)
        else:
            st.info("目前还没有任何记录哦。")


########### 启动方式 ###########
# conda activate tomorrow
# streamlit run app_v2.py
